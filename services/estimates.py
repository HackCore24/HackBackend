from typing import Optional, List

import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from pydantic import BaseModel


class Estimates:
    def __init__(self, project_name, estimate_details):
        self.project_name = project_name
        self.estimate_details = estimate_details
        self.wb = openpyxl.Workbook()
        self.bold_font = Font(bold=True)
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))

    async def header_format(self, ws):
        merge_ranges = ['A6:L6', 'A7:L7', 'A9:L9', 'A10:L10', 'A12:L12', 'A13:L13', 'B21:E21',
                        'A24:A26', 'B24:B26', 'C24:C26', 'D24:D26', 'E24:G24', 'H24:K24', 'L24:M25',
                        'G25:G26', 'H25:H26', 'I25:I26', 'K25:K26']
        center_cells = ['A6', 'A7', 'A9', 'A10', 'A12', 'A13', 'B21']

        for rang in merge_ranges:
            ws.merge_cells(rang)
        for cell in center_cells:
            ws[cell].alignment = self.center_alignment

    async def generate_header(self, ws, chapter):
        header = [
            ["", "СОГЛАСОВАНО:", "", "", "", "", "", "", "", "УТВЕРЖДАЮ:"],
            ["", "", "", "", "", "", "", "", "", ""],
            ["", "_______________", "", "", "", "", "", "", "__________________"],
            ["", '" _____ " ________________ 2024 г.', "", "", "", "", "", "", '"______ " _______________2024 г.'],
            [""],
            [self.project_name],
            ["(наименование стройки)"],
            [""],
            ["ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЕТ № 3"],
            ["(локальная смета)"],
            [""],
            [self.estimate_details],
            ["(наименование работ и затрат, наименование объекта)"],
            [""],
            ["", "Основание:"],
            ["", "Основание", "", chapter.total_price, "тыс. руб."],
            ["", "монтажных работ", "", chapter.total_operation_price, "тыс. руб."],
            ["", "оборудования", "", chapter.total_material_price, "тыс. руб."],
            ["", "Средства на оплату труда", "", chapter.total_work_price, "тыс. руб."],
            [""],
            ["", f"Составлен(а) в базисных ценах по состоянию на 2024"],
            [""],
            [""],
            ["№ пп", "Шифр и номер позиции норматива", "Наименование работ и затрат, единица измерения", "Количество",
             "Стоимость единицы, руб.", "Общая стоимость, руб.",
             "Затраты труда рабочих, чел.-ч, не занятых обслуживанием машин"],
            ["", "", "", "всего", "оплаты труда", "эксплуатации машин", "материалы", "Всего", "оплаты труда",
             "эксплуатация машин", "материалы"],
            ["", "", "", "", "оплаты труда", "в т.ч. оплаты труда", "", "", "", "в т.ч. оплаты труда", "", "на единицу",
             "всего"],
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13],
        ]
        return header

    async def calculate_fields(self, chapters: List[Chapter]):
        chapter = chapters[0]
        for service in chapter.services:
            service.operation_price = service.unit_operation_price * service.quantity
            service.work_price = service.unit_work_price * service.quantity
            service.material_price = service.unit_material_price * service.quantity
            service.total_price = service.operation_price + service.work_price + service.material_price
        chapter.total_price = sum([service.total_price for service in chapter.services])
        chapter.total_material_price = sum([service.material_price for service in chapter.services])
        chapter.total_work_price = sum([service.work_price for service in chapter.services])
        chapter.total_operation_price = sum([service.operation_price for service in chapter.services])
        return chapter

    async def generate_table(self, chapter):
        services = [
            [
                num, service.position_number, service.title, service.quantity, service.unit_work_price,
                service.unit_operation_price, service.unit_material_price, service.total_price, service.work_price,
                service.operation_price, service.material_price
            ]
            for num, service in enumerate(chapter.services)
        ]
        table = [
            [f"Раздел 1. {chapter.title}"],
            *services,
            [f'Итого по разделу 1. {chapter.title}', chapter.total_price]
        ]
        return table

    async def generate_footer(self, chapter: Chapter):
        footer = [
            ["ИТОГИ ПО СМЕТЕ:"],
            ["Итого прямые затраты по смете в ценах 2024г.", "", "", "", "", "", "", str(chapter.total_price), str(chapter.total_work_price), str(chapter.total_operation_price), str(chapter.total_material_price)],
            ["Накладные расходы", "", "", "", "", ""],
            ["Сметная прибыль", "", "", "", "", ""],
            ["Итоги по смете:"],
            ["Итого Монтажные работы", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["Итого Оборудование", "", "", "", "", ""],
            ["Итого", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["Справочно, в ценах 2024г:"],
            ["Материалы", "", "", "", "", ""],
            ["Машины и механизмы", "", "", "", "", ""],
            ["Фонд оплаты труда", "", "", "", "", ""],
            ["Оборудование", "", "", "", "", ""],
            ["Накладные расходы", "", "", "", "", ""],
            ["Сметная прибыль", "", "", "", "", ""],
            ["ВСЕГО по смете", "", "", "", "", "", "", chapter.total_price],
            [""],
            [""],
            ["Составил: ___________________________"],
            ["(должность, подпись, расшифровка)"],
            ["Проверил: ___________________________"],
            ["(должность, подпись, расшифровка)"],
        ]
        return footer

    async def addition(self, ws):

        merge_ranges = ['A28:M28', 'A31:G31', 'A32:M32',
                        'A33:G33', 'A34:G34', 'A35:G35', 'A36:G36', 'A37:G37', 'A38:G38', 'A39:G39',
                        'A40:G40', 'A41:G41', 'A42:G42', 'A43:G43', 'A44:G44', 'A45:G45', 'A46:G46',
                        'A47:G47', 'A50:M50', 'A51:M51', 'A52:M52', 'A53:M53', 'A54:M54']

        center_cells = ["A32", 'A50', "A51", "A52", "A53", "A54"]

        for cell in center_cells:
            ws[cell].alignment = self.center_alignment

        for rng in merge_ranges:
            ws.merge_cells(rng)

        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 34
        ws.column_dimensions['D'].width = 17
        ws.row_dimensions[24].height = 23
        ws.row_dimensions[25].height = 23
        ws.row_dimensions[26].height = 38

        for row in ws.iter_rows(min_row=24, max_row=47, min_col=1, max_col=13):
            for cell in row:
                cell.border = self.thin_border
                cell.font = Font(size=9, name="Arial")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=32, max_row=47, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws['B21'].alignment = Alignment(vertical='center', horizontal=None)
        ws['A28'].alignment = Alignment(vertical='center', horizontal=None)
        ws['A31'].alignment = Alignment(vertical='center', horizontal=None)
        ws['A32'].alignment = Alignment(vertical='center', horizontal="center")
        ws['A9'].font = Font(bold=True, size=12, name="Arial")
        ws['A32'].font = Font(bold=True, size=9, name="Arial")
    async def get_excel(self, chapters: List[Chapter]):
        ws = self.wb.active
        ws.title = "Локальный сметный расчет № 3"

        chapter = await self.calculate_fields(chapters)
        header = await self.generate_header(ws, chapter)
        table = await self.generate_table(chapter)
        footer = await self.generate_footer(chapter)
        rows = [*header, *table, *footer]
        for row in rows:
            ws.append(row)
        await self.header_format(ws)
        await self.addition(ws)
        ws.sheet_view.showGridLines = False
        self.wb.save("example_asd.xlsx")


if __name__ == '__main__':
    class Service(BaseModel):
        position_number: str
        title: str
        quantity: int
        unit_work_price: int
        unit_operation_price: int
        unit_material_price: int
        work_price: Optional[int] = None
        operation_price: Optional[int] = None
        material_price: Optional[int] = None
        total_price: Optional[int] = None


    class Chapter(BaseModel):
        title: str
        total_price: Optional[int] = None
        total_work_price: Optional[int] = None
        total_operation_price: Optional[int] = None
        total_material_price: Optional[int] = None

        services: List[Service]


    service_1 = Service(position_number="TDT 12", title="Прокладка труб, 1000м", quantity=3, unit_work_price=1760,
                        unit_operation_price=782, unit_material_price=120)
    service_2 = Service(position_number="TDT 18.442", title="Установка гидрантов, 1 шт", quantity=2,
                        unit_work_price=2300,
                        unit_operation_price=450, unit_material_price=4000)
    chapter = Chapter(title="Облагораживание територии", services=[service_1, service_2])
    async def main():
        await Estimates(project_name="Тестовый проект", estimate_details="Волшебная смета").get_excel(
            chapters=[chapter])


    import asyncio

    asyncio.run(main())
